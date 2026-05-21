"""
Generate_thesis_Excel.py
========================
Generates a structured Excel overview of all thesis measurements.
One sheet per technology (BT, WLAN, Wired) with metadata extracted
from the folder structure, plus placeholder columns for future
PCAP analysis metrics.

Usage: python Generate_thesis_Excel.py
"""

import os
import re
from pathlib import Path
from pcap_analyzer import analyze_pcap

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install it with:  pip install openpyxl")
    exit(1)


# ============================================================================
#  Configuration
# ============================================================================
BASE_DIR = Path(__file__).resolve().parent
OUTPUT_FILE = BASE_DIR / "Thesis_Measurements_Overview.xlsx"
PCAP_EXTENSIONS = {".pcap", ".pcapng"}


# ============================================================================
#  Column definitions per technology
# ============================================================================
BT_COLUMNS = [
    "#", "Phase", "Cycle Time", "Load Type", "UDP Rate",
    "Capture Point", "Filename", "File Size (MB)", "Relative Path", "Notes",
    # -- Wireless environment (manual input) --
    "Bit Rate", "RSSI (dBm)", "Interference Screenshot",
    # -- Placeholder columns for future PCAP analysis --
    "Avg Latency (ms)", "Max Latency (ms)", "Min Latency (ms)",
    "Jitter (ms)", "Packet Loss (%)",
]

WLAN_COLUMNS = [
    "#", "Phase", "Cycle Time", "Load Type", "UDP Rate", "QoS",
    "Attenuation Line A (dB)", "Attenuation Line B (dB)",
    "Capture Point", "Filename", "File Size (MB)", "Relative Path", "Notes",
    # -- Wireless environment (manual input) --
    "Bit Rate", "RSSI (dBm)", "Interference Screenshot",
    # -- Placeholder columns for future PCAP analysis --
    "Avg Latency (ms)", "Max Latency (ms)", "Min Latency (ms)",
    "Jitter (ms)", "Packet Loss (%)",
]

WIRED_COLUMNS = [
    "#", "PRP Config", "Cable Pulls", "Cycle Time",
    "Capture Point", "Filename", "File Size (MB)", "Relative Path", "Notes",
    # -- Placeholder columns for future PCAP analysis --
    "Avg Latency (ms)", "Max Latency (ms)", "Min Latency (ms)",
    "Jitter (ms)", "Packet Loss (%)",
]

TECH_COLUMNS = {"BT": BT_COLUMNS, "WLAN": WLAN_COLUMNS, "Wired": WIRED_COLUMNS}

# Metric columns get a distinct header color to show they are placeholders
METRIC_COL_NAMES = {
    "Avg Latency (ms)", "Max Latency (ms)", "Min Latency (ms)",
    "Jitter (ms)", "Packet Loss (%)",
}


# ============================================================================
#  Styling
# ============================================================================
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILLS = {
    "BT":    PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid"),
    "WLAN":  PatternFill(start_color="7B2D8B", end_color="7B2D8B", fill_type="solid"),
    "Wired": PatternFill(start_color="E8751A", end_color="E8751A", fill_type="solid"),
}
METRIC_FILL   = PatternFill(start_color="2D6A4F", end_color="2D6A4F", fill_type="solid")
MANUAL_FILL   = PatternFill(start_color="8B6914", end_color="8B6914", fill_type="solid")
ALT_ROW_FILL  = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

# Manual-input columns get a distinct header color (user fills these in)
MANUAL_COL_NAMES = {"Bit Rate", "RSSI (dBm)", "Interference Screenshot"}

DATA_FONT  = Font(name="Calibri", size=10)
NOTES_FONT = Font(name="Calibri", size=10, color="CC0000", bold=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")

# Columns that look better centred
CENTERED_COLS = {
    "#", "Phase", "Cycle Time", "File Size (MB)", "QoS", "Cable Pulls",
    "Capture Point", "UDP Rate", "PRP Config",
    "Attenuation Line A (dB)", "Attenuation Line B (dB)",
    "Load Type",
} | METRIC_COL_NAMES


# ============================================================================
#  Helpers
# ============================================================================
def extract_capture_point(filename: str) -> str:
    """Derive the capture / measurement point from the pcap filename."""
    stem = Path(filename).stem

    if "Between2pairs" in stem:
        return "Between 2 PRP Pairs"
    if "AllLines" in stem:
        return "All Lines"
    if "PRPClient" in stem:
        return "PRP Client"
    if "SwitchAPb" in stem or "SwitchAP" in stem:
        return "Switch AP-b"
    # Careful: match _PRP but NOT NoPRP / SeriePRP
    if re.search(r'(?<!No)(?<!Serie)_PRP', stem):
        return "PRP Client"
    if "End2End" in stem:
        return "End-to-End"

    return ""


def extract_notes(parts: list, filename: str) -> str:
    """Pull warning / error notes from path parts + filename."""
    blob = " ".join(parts) + " " + filename
    notes = []
    if "Watchdog" in blob:
        notes.append("Watchdog triggered")
    if "SwitchError" in blob:
        notes.append("Switch error")
    if "error" in filename.lower() and "SwitchError" not in blob:
        notes.append("Error in capture")
    return "; ".join(notes)


def cycle_time_sort_key(val: str) -> int:
    """Turn '64ms' → 64 for numeric sorting."""
    m = re.search(r'(\d+)', str(val))
    return int(m.group(1)) if m else 9999


# ============================================================================
#  Per-technology path parsers
# ============================================================================
def parse_bt(parts, fname, size_mb, rel):
    """
    Expected paths (parts = folders after 'BT/'):
      P1_Baseline / {cycle} / file
      P2_Load / 5IO_NoUDP   / {cycle} / file
      P2_Load / 5IO_*B_UDP  / {cycle} / {rate} / file
    """
    row = {c: "" for c in BT_COLUMNS}
    row["Filename"], row["File Size (MB)"], row["Relative Path"] = fname, size_mb, rel
    row["Capture Point"] = extract_capture_point(fname)

    if not parts:
        return row

    # Phase
    if "Baseline" in parts[0]:
        row["Phase"] = "Baseline"
        row["Load Type"] = "1 IO"
        if len(parts) >= 2:
            row["Cycle Time"] = parts[1]

    elif "Load" in parts[0]:
        row["Phase"] = "Load"
        if len(parts) >= 2:
            lt = parts[1]
            if "NoUDP"  in lt: row["Load Type"] = "5 IO (No UDP)"
            elif "1400B" in lt: row["Load Type"] = "5 IO + 1400B UDP"
            elif "64B"   in lt: row["Load Type"] = "5 IO + 64B UDP"
        if len(parts) >= 3:
            row["Cycle Time"] = parts[2]
        if len(parts) >= 4:
            rate = parts[3].replace("_Watchdog", "").replace("_SwitchError", "")
            row["UDP Rate"] = rate.replace(",", ".")

    row["Notes"] = extract_notes(parts, fname)

    # PRP Client captures also contain UDP packets when UDP load is present
    if row["Capture Point"] == "PRP Client" and "UDP" in row.get("Load Type", "") and "No UDP" not in row.get("Load Type", ""):
        row["Capture Point"] = "PRP Client + UDP"

    return row


def parse_wlan(parts, fname, size_mb, rel):
    """
    Expected paths (parts = folders after 'WLAN/'):
      P1_Baseline / {cycle} / file
      P2_Load / {QoS|NoQoS} / 5IO_NoUDP   / {cycle} / file
      P2_Load / {QoS|NoQoS} / 5IO_*B_UDP  / {cycle} / file
      P2_Load / {QoS|NoQoS} / 5IO_*B_UDP  / {cycle} / {rate} / file
      P3_Attenuation / 5IO_64ms / LineA_*_LineB_* / file
    """
    row = {c: "" for c in WLAN_COLUMNS}
    row["Filename"], row["File Size (MB)"], row["Relative Path"] = fname, size_mb, rel
    row["Capture Point"] = extract_capture_point(fname)

    if not parts:
        return row

    if "Baseline" in parts[0]:
        row["Phase"] = "Baseline"
        row["Load Type"] = "1 IO"
        if len(parts) >= 2:
            row["Cycle Time"] = parts[1]

    elif "Load" in parts[0]:
        row["Phase"] = "Load"
        if len(parts) >= 2:
            row["QoS"] = "Yes" if parts[1] == "QoS" else "No"
        if len(parts) >= 3:
            lt = parts[2]
            if "NoUDP"  in lt: row["Load Type"] = "5 IO (No UDP)"
            elif "1400B" in lt: row["Load Type"] = "5 IO + 1400B UDP"
            elif "64B"   in lt: row["Load Type"] = "5 IO + 64B UDP"
        if len(parts) >= 4:
            row["Cycle Time"] = parts[3]
        if len(parts) >= 5:
            rate = parts[4].replace("_Watchdog", "").replace("_SwitchError", "")
            row["UDP Rate"] = rate.replace(",", ".")

    elif "Attenuation" in parts[0]:
        row["Phase"] = "Attenuation"
        if len(parts) >= 2:
            m = re.search(r'(\d+ms)', parts[1])
            if m:
                row["Cycle Time"] = m.group(1)
            row["Load Type"] = "5 IO (No UDP)"
        if len(parts) >= 3:
            la = re.search(r'LineA_(-?\d+)', parts[2])
            lb = re.search(r'LineB_(-?\d+)', parts[2])
            if la: row["Attenuation Line A (dB)"] = la.group(1)
            if lb: row["Attenuation Line B (dB)"] = lb.group(1)

    row["Notes"] = extract_notes(parts, fname)

    # PRP Client captures also contain UDP packets when UDP load is present
    if row["Capture Point"] == "PRP Client" and "UDP" in row.get("Load Type", "") and "No UDP" not in row.get("Load Type", ""):
        row["Capture Point"] = "PRP Client + UDP"

    return row


def parse_wired(parts, fname, size_mb, rel):
    """
    Expected paths (parts = folders after 'Wired/'):
      NoPRP / file
      PRP / file
      SeriePRP / {NoCablePulls|CablePulls} / file
    """
    row = {c: "" for c in WIRED_COLUMNS}
    row["Filename"], row["File Size (MB)"], row["Relative Path"] = fname, size_mb, rel
    row["Capture Point"] = extract_capture_point(fname)

    # Cycle time from filename (e.g. Wire_64ms_...)
    m = re.search(r'(\d+ms)', fname)
    if m:
        row["Cycle Time"] = m.group(1)

    if not parts:
        return row

    prp = parts[0]
    if prp == "NoPRP":
        row["PRP Config"] = "No PRP"
    elif prp == "PRP":
        row["PRP Config"] = "PRP"
    elif prp == "SeriePRP":
        row["PRP Config"] = "Series PRP"
        if len(parts) >= 2:
            row["Cable Pulls"] = "Yes" if parts[1] == "CablePulls" else "No"

    row["Notes"] = extract_notes(parts, fname)
    return row


PARSERS = {"BT": parse_bt, "WLAN": parse_wlan, "Wired": parse_wired}


# ============================================================================
#  Sorting helpers  (Phase order → Cycle time → Load type → Rate → Capture)
# ============================================================================
PHASE_ORDER  = {"Baseline": 0, "Load": 1, "Attenuation": 2, "": 99}
LOAD_ORDER   = {"5 IO (No UDP)": 0, "5 IO + 64B UDP": 1, "5 IO + 1400B UDP": 2, "": 99}
PRP_ORDER    = {"No PRP": 0, "PRP": 1, "Series PRP": 2, "": 99}

def sort_key(row):
    return (
        PHASE_ORDER.get(row.get("Phase", ""), 99),
        PRP_ORDER.get(row.get("PRP Config", ""), 99),
        row.get("Cable Pulls", ""),
        cycle_time_sort_key(row.get("Cycle Time", "")),
        LOAD_ORDER.get(row.get("Load Type", ""), 99),
        row.get("QoS", ""),
        cycle_time_sort_key(row.get("UDP Rate", "")),
        row.get("Attenuation Line A (dB)", ""),
        row.get("Capture Point", ""),
    )


# ============================================================================
#  Walk & collect
# ============================================================================
def collect_measurements(tech: str) -> list:
    """Walk a technology folder and return a list of parsed row dicts."""
    tech_dir = BASE_DIR / tech
    if not tech_dir.exists():
        print(f"  WARNING: {tech_dir} not found - skipping.")
        return []

    parser = PARSERS[tech]
    rows = []

    for root, _dirs, files in os.walk(tech_dir):
        for fname in sorted(files):
            if Path(fname).suffix.lower() not in PCAP_EXTENSIONS:
                continue

            full = Path(root) / fname
            size_mb = round(full.stat().st_size / (1024 * 1024), 2)

            rel = full.relative_to(tech_dir)
            folder_parts = list(rel.parent.parts)      # folders only (no filename)
            rel_str = str(rel).replace("\\", "/")       # clean path string

            row_dict = parser(folder_parts, fname, size_mb, rel_str)
            
            cycle_time = row_dict.get("Cycle Time", "64ms")
            try:
                digits = "".join(filter(str.isdigit, str(cycle_time)))
                ct_val = int(digits) if digits else 64
            except ValueError:
                ct_val = 64
                
            metrics = analyze_pcap(str(full), cycle_time_ms=ct_val)
            for k in ["Avg Latency (ms)", "Max Latency (ms)", "Min Latency (ms)", "Jitter (ms)", "Packet Loss (%)"]:
                if k in metrics:
                    row_dict[k] = metrics.get(k, "")
                    
            rows.append(row_dict)

    rows.sort(key=sort_key)
    return rows


# ============================================================================
#  Excel generation
# ============================================================================
def write_excel(all_data: dict):
    wb = Workbook()
    wb.remove(wb.active)                                # drop default sheet

    for tech in ("BT", "WLAN", "Wired"):
        rows = all_data.get(tech, [])
        columns = TECH_COLUMNS[tech]
        fill = HEADER_FILLS[tech]

        ws = wb.create_sheet(title=tech)

        # ── Header row ────────────────────────────────────────────────────
        for ci, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=ci, value=col_name)
            cell.font      = HEADER_FONT
            cell.alignment  = CENTER
            cell.border     = THIN_BORDER
            if col_name in METRIC_COL_NAMES:
                cell.fill = METRIC_FILL
            elif col_name in MANUAL_COL_NAMES:
                cell.fill = MANUAL_FILL
            else:
                cell.fill = fill

        # ── Data rows ─────────────────────────────────────────────────────
        for ri, row_data in enumerate(rows, 2):
            row_data["#"] = ri - 1                      # auto-number

            for ci, col_name in enumerate(columns, 1):
                val = row_data.get(col_name, "")
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border    = THIN_BORDER
                cell.font      = NOTES_FONT if (col_name == "Notes" and val) else DATA_FONT
                cell.alignment = CENTER if col_name in CENTERED_COLS else LEFT

                if ri % 2 == 0:
                    cell.fill = ALT_ROW_FILL

        # ── Column widths (auto-fit) ──────────────────────────────────────
        for ci, col_name in enumerate(columns, 1):
            max_len = len(col_name)
            for ri in range(2, len(rows) + 2):
                v = ws.cell(row=ri, column=ci).value
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 65)

        # ── Freeze header + auto-filter ───────────────────────────────────
        ws.freeze_panes = "A2"
        last_col = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A1:{last_col}{max(len(rows) + 1, 2)}"

        print(f"  {tech:6s} -> {len(rows):3d} pcap files")

    wb.save(OUTPUT_FILE)
    print(f"\nDone! Excel saved to: {OUTPUT_FILE}")


# ============================================================================
#  Main
# ============================================================================
def main():
    print(f"Base directory : {BASE_DIR}")
    print(f"Output file    : {OUTPUT_FILE}\n")
    print("Scanning for .pcap / .pcapng files ...")

    data = {}
    for tech in ("BT", "WLAN", "Wired"):
        data[tech] = collect_measurements(tech)

    print("\nGenerating Excel ...")
    write_excel(data)
    print("Done!")


if __name__ == "__main__":
    main()